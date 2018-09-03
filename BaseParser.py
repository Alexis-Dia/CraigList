from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
import csv
import sys
import re
import os
import wmi
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime
from selenium.webdriver.common.action_chains import ActionChains
import datetime
from openpyxl import Workbook, load_workbook

phantomjs_path = "D:\phantomjs.exe"
chromedriver_path = "D:\chromedriver.exe"
url = 'https://www.ktmdealer.net/Shop/'
login = 'armin_brem'
password = 'armin_153'
listOfCountries = ['Canada', 'USA']

def getNumberOfElements(driver, pattern):
    numbersOfItems = -1
    try:
       numbersOfItems = (len(driver.find_elements_by_xpath(pattern)))
    except:
        print('Err in contactName1 ...')
    return numbersOfItems

def ctreatePattern(pattern, numberOfItem):
    return pattern + str(numberOfItem) + ']/a'

def ctreatePattern2(pattern, numberOfItem):
    return '//li[' + str(numberOfItem) + ']' + pattern

def getHref(driver, pattern):
    href = driver.find_element_by_xpath(pattern).get_attribute("href")
    return href

def clickOnBulletPoint(driver, pattern):
    driver.get(pattern)

def getAllHrefs(driver, numbersOfItems, pattern1):
    hrefs = []
    for i in range(1, numbersOfItems + 1):
        pattern = ctreatePattern(pattern1, i)
        hrefs.append(getHref(driver, pattern))
    return hrefs

def getAllHrefs2(driver, numbersOfItems, pattern1):
    hrefs = []
    for i in range(1, numbersOfItems + 1):
        pattern = ctreatePattern2(pattern1, i)
        hrefs.append(getHref(driver, pattern))
    return hrefs

def parseCitiesUrl(driver, pattern2):
    numbersOfCities = int(getNumberOfElements(driver, pattern2))
    hrefsOfCities = []
    for i in range(1, numbersOfCities + 1):
        pattern = ctreatePattern(pattern2 + '[', i)
        hrefsOfCities.append(getHref(driver, pattern))
    return hrefsOfCities

def parseJobsUrl(driver, pattern, numberOfElements):
    hrefsOfJobs = []
    for i in range(1, numberOfElements + 1):
        pattern = ctreatePattern2(pattern, i)
        hrefsOfJobs.append(getHref(driver, pattern))
    return hrefsOfJobs

def clickOnAllBulletPoints(driver):
    pattern1 = 'html/body/div/section/div/ul/li'
    pattern2 = ".//*[@id='rightbar']/ul/li[1]/ul/li"
    numbersOfItems = int(getNumberOfElements(driver, pattern1))
    hrefs = getAllHrefs(driver, numbersOfItems, pattern1 + '[')
    hrefsOfCitiesList = []
    for i in hrefs:
        clickOnBulletPoint(driver, i)
        hrefsOfCities1 = parseCitiesUrl(driver, pattern2)
        hrefsOfCitiesList = hrefsOfCitiesList + hrefsOfCities1
    return hrefsOfCitiesList

def getAllHrefsData(driver):
    pattern = "//a[contains(@class, 'result-title hdrlnk')]"
    numbersOfItems = int(getNumberOfElements(driver, pattern))
    hrefsOfJobsList = getAllHrefs2(driver, numbersOfItems, pattern)
    return hrefsOfJobsList

def getHrefsJobs(driver, keyWords, param1, hrefsOfCitiesSet):
    hrefsOfJobsListAll = []
    for i in hrefsOfCitiesSet:
        for j in keyWords:
            url = str(i) + 'd/' + str(param1) + '?query=' + str(j) + '&employment_type=2'
            driver.get(url)
            hrefsOfJobsList = getAllHrefsData(driver)
            hrefsOfJobsListAll = hrefsOfJobsListAll + hrefsOfJobsList
    setOfHrefsJobsSoftwareDbQa = set(hrefsOfJobsListAll)
    print(setOfHrefsJobsSoftwareDbQa)
    return setOfHrefsJobsSoftwareDbQa

def getUrls(listOfStates, driver, baseUrl, *args):
    hrefsOfCities = []
    for i in listOfStates:
        print(baseUrl + i[1])
        driver.get(baseUrl + i[1])
        hrefsOfCities = hrefsOfCities + clickOnAllBulletPoints(driver)
    if len(args) == 1:
        hrefsOfCities = hrefsOfCities + args[0]
    setOfHrefsJobsSoftwareDbQa = set(hrefsOfCities)
    #setOfHrefsJobsSoftwareDbQa.update(setOfHrefsRegion)
    return setOfHrefsJobsSoftwareDbQa

def getCurrentDate():
    now = datetime.datetime.now()
    return str(now.year) + '-' + str(now.month) + '-' + str(now.day) + '-'

def keyWordsBuilder(keyWords):
    line = '-'
    for i in keyWords:
        # if i == '*':
        #     line += '*' + '-'
        #     line = line.replace('*', '(all)')
        # else:
        #     line += i + '-'
        line += i + '-'
    line = line.replace('*', '(all)')
    return line

def writeResToFile(driver, param, setOfJobsHrefsSoftwareDbQa, keyWords):
    date = getCurrentDate()
    fileName = 'results/' + date + param.replace('/', '-') + keyWordsBuilder(keyWords)
    createFile(fileName)
    for i in setOfJobsHrefsSoftwareDbQa:
        driver.get(i)
        title = driver.find_element_by_xpath(".//*[@id='titletextonly']").text
        body = driver.find_element_by_xpath(".//*[@id='postingbody']").text
        savePageData(fileName, title, body, i)

# def remove_duplicates(list):
#     return list(set(list))
#
# def logging(driver, url, login, password):
#     driver.get(url)
#     element = driver.find_element_by_xpath("//*[@id='UserName']")
#     element.send_keys(login)
#     element = driver.find_element_by_xpath("//*[@id='Password']")
#     element.send_keys(password)
#     element = driver.find_element_by_xpath("//*[@id='SignInForm']/fieldset/input")
#     element.click()
#     time.sleep(5)
#
# def saveToXlsx(driver, href, fileName):
#
#     fileName = fileName + '.xlsx'
#
#     wb = Workbook()
#     wb.save(filename = fileName)
#
#     wb = load_workbook(filename = fileName)
#
#     # grab the active worksheet
#     ws = wb.active
#
#     # Data can be assigned directly to cells
#     #ws['A1'] = 42
#
#     # Rows can also be appended
#     ws.append([1, 2, 3])
#
#     # Python types will automatically be converted
#     #ws['A2'] = datetime.datetime.now()
#
#     # Save the file
#     wb.save(fileName + ".xlsx")
#

def createFile(fileName):
    fileName = fileName + '.xlsx'
    wb = Workbook()
    wb.save(filename = fileName)
#
def savePageData(fileName, title, data, href):

    fileName = fileName + '.xlsx'

    wb = load_workbook(filename = fileName)
    ws = wb.active
    ws.append([title, data, href])
    wb.save(fileName)

def readFile(fileName):
    fileName = fileName + '.xlsx'
    setOfHrefsJobsSoftwareDbQa = []
    wb = load_workbook(filename=fileName)
    ws = wb["Sheet"]
    for cell in ws['A']:
        if cell.value is not None:
            setOfHrefsJobsSoftwareDbQa.append(str(cell.value))
    return set(setOfHrefsJobsSoftwareDbQa)

#
# def saveAllData(catalogueName, data):
#
#     fileName = 'IC GC ' + catalogueName + '.xlsx'
#
#     wb = load_workbook(filename = fileName)
#     ws = wb.active
#     ws.append(data)
#     wb.save(fileName)
#
# def clickOnCatalogue(driver, cataloqueName):
#     print('1')
#     fileName = 'Hrefs' + cataloqueName
#
#     if ((cataloqueName == 'Retail Systems') or (cataloqueName == 'Marketing Material')  or (cataloqueName == 'Spezialwerkzeuge') or (cataloqueName == 'PowerParts')):
#         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#         time.sleep(1.5)
#
#     element = driver.find_element_by_xpath(".//span[contains(@class, 'innerSlideCaption')][text()='" + cataloqueName + "']/../..//div[1]/img")
#     element.click()
#     time.sleep(3)
#     print('1.1')
#     # element = driver.find_element_by_xpath(".//div[contains(@class, 'divDetailRow row product-list-model')][12]//img")
#     # print('1.1')
#     # ActionChains(driver).move_to_element(element).perform()
#     # print('1.2')
#     time.sleep(3)
#
#     numberOfProducts = 0
#     try:
#         print('1.2')
#         element = driver.find_element_by_xpath("//div[contains(@class, 'row shop-top-header')]/div[1]/span")
#         numberOfProducts = element.text
#         numberOfProducts = re.findall(r'\d+', numberOfProducts)
#         numberOfProducts = int(numberOfProducts[0])
#         print('1.3')
#     except:
#         numberOfProducts = 3000
#
#     print('1.4')
#     print('numberOfProducts: ' + str(numberOfProducts))
#     createFile(fileName)
#
#     index = 1;
#     while index != numberOfProducts:
#         href = ""
#         artickle = ""
#         title = ""
#         try:
#             time.sleep(1.5)
#             inputElement = driver.find_element_by_xpath("//div[contains(@class, 'divDetailRow row product-list-model')][" + str(index) + "]//div[contains(@class, 'model-title')]//a")
#             href = inputElement.get_attribute("href")
#
#             element = driver.find_element_by_xpath("//div[@id='product-list-container']//descendant::div[@id='divModelDetail'][1]//div[@class='divDetailRow row product-list-model'][" + str(index) + "]//div[contains(@class,'small-title')]/span")
#             artickle = element.text
#
#             element = driver.find_element_by_xpath("//div[contains(@class, 'divDetailRow row product-list-model')][" + str(index) + "]//div[contains(@class, 'model-title')]//a")
#             title = element.text
#
#             saveAllHrefs(href, fileName, artickle, title)
#         except:
#             href = ""
#             artickle = ""
#             title = ""
#
#         print(index, " ", artickle, " ", title, " ", href)
#         index = index + 1
#         if index % 12 == 0:
#             driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#             time.sleep(15)
#
# def backToMainPage(driver, url):
#     driver.get(url)
#
# def parseXlsx(cataloqueName):
#     fileName = 'Hrefs' + cataloqueName + '.xlsx'
#     wb = load_workbook(filename = fileName)
#     sheet_ranges = wb['Sheet']
#     numbersOfRows = sheet_ranges.max_row
#     print(numbersOfRows)
#
#     myList = []
#     for x in range(1, numbersOfRows + 1):
#         href = sheet_ranges['A' + str(x)].value
#         title = sheet_ranges['B' + str(x)].value
#         artickle = sheet_ranges['C' + str(x)].value
#         details = ""
#         variants = ""
#         access = ""
#         items = [artickle, title, details, variants, access, href]
#         #print(x, " ", href)
#         myList.append(items)
#
#     return myList
#
# def parseXlsxForIcGc(cataloqueName):
#     fileName = 'Hrefs-' + cataloqueName + '.xlsx'
#     wb = load_workbook(filename = fileName)
#     sheet_ranges = wb['Sheet']
#     numbersOfRows = sheet_ranges.max_row
#     print(numbersOfRows)
#
#     myList = []
#     for x in range(2, numbersOfRows + 1):
#         cataloqueName = sheet_ranges['A' + str(x)].value
#         subCataloqueName = sheet_ranges['B' + str(x)].value
#         href = sheet_ranges['C' + str(x)].value
#         items = [cataloqueName, subCataloqueName, href]
#         print(cataloqueName, " ", subCataloqueName, " ", href, " ")
#         myList.append(items)
#
#     return myList
#
# def parsePage(driver, filename, href):
#     if href != None:
#         driver.get(href)
#
# def addDescToXlsx(driver, list, fileName):
#     fileName = 'IC GC ' + fileName;
#     createFile(fileName)
#     for i in range(len(list)):
#         href = list[i][2]
#         if href != None:
#             print(href)
#             try:
#                 driver.get(href)
#             except:
#                 try:
#                     time.sleep(10)
#                     print('eeeeeeerrrrrrrrooooorrrrrr1 in get')
#                     driver.get(href)
#                 except:
#                     try:
#                         time.sleep(10)
#                         print('eeeeeeerrrrrrrrooooorrrrrr2 in get')
#                         driver.get(href)
#                     except:
#                         try:
#                             time.sleep(10)
#                             print('eeeeeeerrrrrrrrooooorrrrrr3 in get')
#                             driver.get(href)
#                         except:
#                             continue
#             companyName = ""
#             try:
#                 element = driver.find_element_by_xpath("html/body/main/div[1]/header/h2")
#                 companyName = element.text
#             except:
#                 print('Err in companyName ...')
#                 companyName = ""
#
#                 # try:
#                 #     element = driver.find_element_by_xpath("//div[contains(text(),'Legal/Operating Name')]/..//div[2]")
#                 #     companyName = element.text
#                 # except:
#                 #     try:
#                 #         element = driver.find_element_by_xpath("//div[contains(text(),'Legal Name')]/..//div[2]")
#                 #         companyName = element.text
#                 #     except:
#                 #         try:
#                 #             element = driver.find_element_by_xpath("//div[contains(text(),'Operating Name')]/..//div[2]")
#                 #             companyName = element.text
#                 #         except:
#                 #             print('Err in companyName ...')
#                 #             companyName = ""
#
#             website = ""
#             try:
#                 inputElement = driver.find_element_by_xpath(".//span[contains(@class, 'fa fa-globe mrgn-rght-sm')]/../../div[2]/a")
#                 website = inputElement.get_attribute("href")
#             except:
#                 print('Err in website ...')
#                 website = ""
#
#             email = ""
#             try:
#                 inputElement = driver.find_element_by_xpath(".//span[contains(@class, 'fa fa-envelope mrgn-rght-sm')]/../../div[2]/a")
#                 email = inputElement.text.strip()
#             except:
#                 print('Err in email ...')
#                 email = ""
#
#             phone = ""
#             try:
#                 inputElement = driver.find_element_by_xpath(".//span[contains(@class, 'fa fa-phone mrgn-rght-sm')]/../../div[2]/a")
#                 phone = inputElement.text.strip()
#             except:
#                 print('Err in phone ...')
#                 phone = ""
#
#             adress = ""
#             try:
#                 inputElement = driver.find_element_by_xpath("//h2[contains(text(),'Location address')]/../../div")
#                 adressArr = inputElement.text.splitlines()
#                 adress = adressArr[1].strip()
#             except:
#                 try:
#                     inputElement = driver.find_element_by_xpath("//h2[contains(text(),'Location address')]/../../div")
#                     adressArr = inputElement.text.splitlines()
#                     adress = adressArr[0].strip()
#                 except:
#                     print('Err in adress ...')
#                     adress = ""
#
#             driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#             #time.sleep(0.1)
#
#             numbersOfItems = -1
#             try:
#                 numbersOfItems = (len(driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-3')]/strong")))
#             except:
#                 print('Err in contactName1 ...')
#                 numbersOfItems = ""
#
#             contactName1 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-3')]/strong")
#                 contactName1 = inputElements[0].text.strip()
#             except:
#                 print('Err in contactName1 ...')
#                 contactName1 = ""
#
#             contactTitle1 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Title')]/../../div[2]")
#                 contactTitle1 = inputElements[0].text.strip()
#             except:
#                 print('Err in contactTitle1 ...')
#                 contactTitle1 = ""
#
#             areaOfResponsibility1 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Area of Responsibility')]/../../div[2]")
#                 areaOfResponsibility1 = inputElements[0].text.strip()
#             except:
#                 print('Err in areaOfResponsibility1 ...')
#                 areaOfResponsibility1 = ""
#
#             telephone1 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Telephone')]/../../div[2]")
#                 telephone1 = inputElements[0].text.strip()
#             except:
#                 print('Err in telephone1 ...')
#                 telephone1 = ""
#
#             email1 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Email')]/../../div[2]")
#                 email1 = inputElements[0].text.strip()
#             except:
#                 print('Err in email1 ...')
#                 email1 = ""
#
#
#             contactName2 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-3')]/strong")
#                 contactName2 = inputElements[1].text.strip()
#             except:
#                 print('Err in contactName2 ...')
#                 contactName2 = ""
#
#             contactTitle2 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Title')]/../../div[2]")
#                 contactTitle2 = inputElements[1].text.strip()
#             except:
#                 print('Err in contactTitle2 ...')
#                 contactTitle2 = ""
#
#             areaOfResponsibility2 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Area of Responsibility')]/../../div[2]")
#                 areaOfResponsibility2 = inputElements[1].text.strip()
#             except:
#                 print('Err in areaOfResponsibility2 ...')
#                 areaOfResponsibility2 = ""
#
#             telephone2 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Telephone')]/../../div[2]")
#                 telephone2 = inputElements[1].text.strip()
#             except:
#                 print('Err in telephone2 ...')
#                 telephone2 = ""
#
#             email2 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Email')]/../../div[2]")
#                 email2 = inputElements[1].text.strip()
#             except:
#                 print('Err in email2 ...')
#                 email2 = ""
#
#
#             contactName3 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-3')]/strong")
#                 contactName3 = inputElements[2].text.strip()
#             except:
#                 print('Err in contactName3 ...')
#                 contactName3 = ""
#
#             contactTitle3 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Title')]/../../div[2]")
#                 contactTitle3 = inputElements[2].text.strip()
#             except:
#                 print('Err in contactTitle3 ...')
#                 contactTitle3 = ""
#
#             areaOfResponsibility3 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Area of Responsibility')]/../../div[2]")
#                 areaOfResponsibility3 = inputElements[2].text.strip()
#             except:
#                 print('Err in areaOfResponsibility3 ...')
#                 areaOfResponsibility3 = ""
#
#             telephone3 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Telephone')]/../../div[2]")
#                 telephone3 = inputElements[2].text.strip()
#             except:
#                 print('Err in telephone3 ...')
#                 telephone3 = ""
#
#             email3 = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//*[@id='details-panel1']/div/section[1]/div/div[contains(@class, 'col-md-5')]/strong[contains(text(),'Email')]/../../div[2]")
#                 email3 = inputElements[2].text.strip()
#             except:
#                 print('Err in email3 ...')
#                 email3 = ""
#
#             numberOfEmployees = ""
#             try:
#                 inputElements = driver.find_elements_by_xpath(".//strong[contains(text(),'Number of Employees')]/../../div[2]")
#                 numberOfEmployees = inputElements[0].text.strip()
#             except:
#                 print('Err in numberOfEmployees ...')
#                 numberOfEmployees = ""
#
#             catalogueName = list[i][0]
#             subCatalogueName = list[i][1]
#
#             data = [catalogueName, subCatalogueName, companyName, website, email, phone, adress,
#                     contactName1, contactTitle1, areaOfResponsibility1, telephone1, email1,
#                     contactName2, contactTitle2, areaOfResponsibility2, telephone2, email2,
#                     contactName3, contactTitle3, areaOfResponsibility3, telephone3, email3,
#                     numberOfEmployees]
#             try:
#                 saveAllData(catalogueName, data)
#             except:
#                 continue
#
# def parsePersonalPage(driver, href, catalogueName):
#     driver.get(href)
#
#     companyName = ""
#     try:
#         element = driver.find_element_by_xpath(".//*[@id='containerContent']/div/div/h2")
#         companyName = element.text
#     except:
#         print('Err in companyName ...')
#         companyName = ""
#
#     line2 = ""
#     try:
#         element = driver.find_element_by_xpath(".//*[@id='containerContent']/div/div/div[1]/p")
#         line2 = element.text.splitlines()[1].strip()
#     except:
#         print('Err in line3 ...')
#         line2 = ""
#
#     line3 = ""
#     try:
#         element = driver.find_element_by_xpath(".//*[@id='containerContent']/div/div/div[1]/p")
#         line3 = element.text.splitlines()[2].strip()
#     except:
#         print('Err in line3 ...')
#         line3 = ""
#
#     city = line2
#     country = line3
#
#     if (line3 == ""):
#         for k in range(len(listOfCountries)):
#             if listOfCountries[k] == line2:
#                 country = line2
#                 city = ""
#
#     for i in range(len(listOfCountries)):
#         if listOfCountries[i] in line3:
#             if len(line3) != len(listOfCountries[i]):
#                 city = line3.replace(listOfCountries[i], '')
#                 country = listOfCountries[i]
#                 break
#
#     website = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileWebsite')]/a")
#         website = inputElement.get_attribute("href")
#     except:
#         print('Err in website ...')
#         website = ""
#
#     email = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileEmail')]/a")
#         email = inputElement.text.strip()
#     except:
#         print('Err in email ...')
#         email = ""
#
#     phone1 = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//ul[contains(@class, 'profileNumbers')]/li[1]")
#         phone1 = inputElement.text.strip().replace('Tel:', '').strip()
#     except:
#         print('Err in phone1 ...')
#         phone1 = ""
#
#     phone2 = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//ul[contains(@class, 'profileNumbers')]/li[2]")
#         phone2 = inputElement.text.strip()
#     except:
#         print('Err in phone2 ...')
#         phone2= ""
#
#     profileContact1 = ""
#     profileContactTitle1 = ""
#     profileContactName1 = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileContact1')]")
#         profileContactList = inputElement.text.strip().split(":")
#         profileContactTitle1 = profileContactList[0].strip()
#         profileContactName1 = profileContactList[1].strip()
#
#         if profileContactName1 == "":
#             profileContactTitle1 = ""
#     except:
#         print('Err in profileContact1 ...')
#         profileContact1 = ""
#         profileContactTitle1 = ""
#         profileContactName1 = ""
#
#     profileContact2 = ""
#     profileContactTitle2 = ""
#     profileContactName2 = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileContact2')]")
#         profileContactList = inputElement.text.strip().split(":")
#         profileContactTitle2 = profileContactList[0].strip()
#         profileContactName2 = profileContactList[1].strip()
#
#         if profileContactName2 == "":
#             profileContactTitle2 = ""
#     except:
#         print('Err in profileContact2 ...')
#         profileContact2 = ""
#         profileContactTitle2 = ""
#         profileContactName2 = ""
#
#     profileContact3 = ""
#     profileContactTitle3 = ""
#     profileContactName3 = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileContact3')]")
#         profileContactList = inputElement.text.strip().split(":")
#         profileContactTitle3 = profileContactList[0].strip()
#         profileContactName3 = profileContactList[1].strip()
#
#         if profileContactName3 == "":
#             profileContactTitle3 = ""
#     except:
#         print('Err in profileContact3 ...')
#         profileContact3 = ""
#         profileContactTitle3 = ""
#         profileContactName3 = ""
#
#     profileContact4 = ""
#     profileContactTitle4 = ""
#     profileContactName4 = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileContact4')]")
#         profileContactList = inputElement.text.strip().split(":")
#         profileContactTitle4 = profileContactList[0].strip()
#         profileContactName4 = profileContactList[1].strip()
#
#         if profileContactName4 == "":
#             profileContactTitle4 = ""
#     except:
#         print('Err in profileContact4 ...')
#         profileContact4 = ""
#         profileContactTitle4 = ""
#         profileContactName4 = ""
#
#     profileFoundedEst = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileFounded')]")
#         profileFoundedList = inputElement.text.strip().split(":")
#         profileFoundedEst = profileFoundedList[1].strip()
#
#     except:
#         print('Err in profileFounded ...')
#         profileFoundedEst = ""
#
#     profileRevenueRev = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileRevenue')]")
#         profileRevenueList = inputElement.text.strip().split(":")
#         profileRevenueRev = profileRevenueList[1].strip()
#
#     except:
#         print('Err in profileRevenue ...')
#         profileRevenueRev = ""
#
#     profileStaff = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//li[contains(@class, 'profileStaff')]")
#         profileStaffList = inputElement.text.strip().split(":")
#         profileStaff = profileStaffList[1].strip()
#
#     except:
#         print('Err in profileFounded ...')
#         profileStaff = ""
#
#     specialtyFin = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//*[@id='containerContent']/div/div/div[2]")
#         specialty1 = inputElement.text
#         specialtyFin = ", ".join(specialty1.split("\n"))
#
#     except:
#         print('Err in specialty1 ...')
#         specialtyFin = ""
#
#         profileSummary = ""
#     try:
#         inputElement = driver.find_element_by_xpath(".//div[contains(@class, 'profileSummary')]")
#         profileSummary = inputElement.text
#
#     except:
#         print('Err in profileSummary ...')
#         profileSummary = ""
#
#     data = [catalogueName, companyName, city, country, website, email, phone1, phone2, profileContactTitle1, profileContactName1, profileContactTitle2, profileContactName2, profileContactTitle3, profileContactName3, profileContactTitle4, profileContactName4, profileFoundedEst, profileRevenueRev, profileStaff, specialtyFin, profileSummary]
#     saveAllData(catalogueName, data)
#
# def parse(driver, base_url, catalogueName):
#
#     createFile(catalogueName)
#
#     driver.get(base_url);
#
#     driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
#     element = driver.find_element_by_xpath(".//p/input[contains(@class, 'buttonEnhanced')]")
#     element.click()
#     driver.get(base_url);
#     time.sleep(15)
#
#     numbersOfItems = (len(driver.find_elements_by_xpath(".//ul[contains(@class, 'listResults')]/li")))
#     print(numbersOfItems)
#
#     listOfhrefs = []
#     for index in range(1, numbersOfItems + 1):
#         inputElement = driver.find_element_by_xpath(".//ul[contains(@class, 'listResults')]/li[" + str(index) + "]/ul/li/a")
#         href = inputElement.get_attribute("href")
#         listOfhrefs.append(href)
#         print(href)
#
#     for index in range(len(listOfhrefs)):
#         href = listOfhrefs[index]
#         if href != None:
#             print(href)
#             parsePersonalPage(driver, href, catalogueName)
#
# def getListOfSubCatalogueItems(driver, subHref, filename, catalogueName, subCatalogueName):
#     driver.get(subHref)
#     numbersOfItems = (len(driver.find_elements_by_xpath("html/body/main/ul/li")))
#     for index in range(1, numbersOfItems + 1):
#         inputElement = driver.find_element_by_xpath("html/body/main/ul/li[" + str(index) + "]/a")
#         hrefOfItem = inputElement.get_attribute("href")
#         print(catalogueName + " " + subCatalogueName + " " + hrefOfItem)
#         data = [catalogueName, subCatalogueName, hrefOfItem]
#         saveAllData(filename, data)
#
# def parseListOfCataloque(driver, catalogueName, listOfSubCataloque):
#     filename = 'Hrefs-' + catalogueName
#     for i in range(len(listOfSubCataloque)):
#         subCataloque = listOfSubCataloque[i][0]
#         if subCataloque != None:
#             driver.get(listOfSubCataloque[i][0])
#             getListOfSubCatalogueItems(driver, listOfSubCataloque[i][0], filename, catalogueName, listOfSubCataloque[i][2])
#
# def getListOfCataloque(driver, base_url, catalogueName):
#     filename = 'Hrefs-' + catalogueName
#     createFile(filename)
#     driver.get(base_url)
#     listOfCataloque = []
#     numbersOfItems = (len(driver.find_elements_by_xpath(".//*[@id='NaicsTbl']/tbody/tr")))
#     for index in range(2, numbersOfItems + 1):
#         inputElement = driver.find_element_by_xpath(".//*[@id='NaicsTbl']/tbody/tr[" + str(index) + "]/td[2]/a")
#         subHref = inputElement.get_attribute("href")
#         inputElement = driver.find_element_by_xpath(".//*[@id='NaicsTbl']/tbody/tr[" + str(index) + "]/td[2]/a")
#         subCatalogueName = inputElement.text
#         listOfCataloque.append([subHref, catalogueName, subCatalogueName])
#     return listOfCataloque
#
# def parseCataloque(driver, base_url, catalogueName):
#     listOfCataloque = getListOfCataloque(driver, base_url, catalogueName)
#     parseListOfCataloque(driver, catalogueName, listOfCataloque)
#
# def removeAndSortRepeatedElements(listFromFile, fileName):
#     newList = []
#     newList2 = []
#
#     print(len(listFromFile))
#
#     for i in range(0, len(listFromFile)):
#         #print(listFromFile[i][2])
#         #item = listFromFile[i][2]
#         #item = listFromFile[i][0] + delimeter + listFromFile[i][2]
#         #item = listFromFile[i][0] + delimeter + listFromFile[i][1] + delimeter + listFromFile[i][2]
#         item = Href(listFromFile[i][0], listFromFile[i][1], listFromFile[i][2])
#         newList.append(item)
#
#     print(len(newList))
#
#     setOfUniqueHrefs = set()
#     setOfUniqueHrefs2 = set()
#     for i in range(0, len(newList)):
#         if newList[i].subCatalogueName != fileName:
#             setOfUniqueHrefs.add(newList[i].href)
#             #print("------      ", newList[i].catalogueName, " ", newList[i].subCatalogueName, " ", newList[i].href)
#
#     print("setOfUniqueHrefs = " + str(len(setOfUniqueHrefs)))
#
#     # for i in range(0, len(newList)):
#     #     uniqueIndex = -1;
#     #     for index, val in enumerate(setOfUniqueHrefs):
#     #         if newList[i].href == val:
#     #             newList2.append(newList[i])
#     #             break
#     #     if len(newList2) == len(setOfUniqueHrefs):
#     #         break
#
#     for index, val in enumerate(setOfUniqueHrefs):
#         uniqueIndex = -1;
#         for i in range(0, len(newList)):
#             if newList[i].href == val:
#                 if newList[i].subCatalogueName != fileName:
#                     newList2.append(newList[i])
#                     break
#         if len(newList2) == len(setOfUniqueHrefs):
#             break
#
#     print("newList2 = " + str(len(newList2)))
#     newList2.sort(key = lambda c: c.subCatalogueName)
#
#     for i in range(0, len(newList2)):
#         print(newList2[i].catalogueName, " ", newList2[i].subCatalogueName, " ", newList2[i].href)
#         #print(newList2[i].href)
#
#     #addDescToXlsx(driver, list, fileName)
#
#     return newList2
#
# # def moveToElement(driver):
# #     element = driver.find_element_by_xpath(".//*[@id='jobs-informer']/div[2]")
# #     while True:
# #         driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
